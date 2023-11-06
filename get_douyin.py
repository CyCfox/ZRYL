import requests
from bs4 import BeautifulSoup
import re
import math
from openpyxl import Workbook
import jieba
import matplotlib.pyplot as plt
from wordcloud import WordCloud
from PIL import Image
import numpy as np

barrages_num=20#单个视频爬取弹幕数
video_num=30#爬取视频数

# 发送GET请求，获取视频页面的HTML内容
def get_video_html(url):
    response = requests.get(url)
    return response.text

# 发送GET请求，获取B站搜索结果页面的HTML内容
def get_search_results_html():
    #设计html标头
    page,html=1,''#page变量用于自动翻页,html用于存储所有网页的html文件
    url = "https://search.bilibili.com/all?keyword=抖音"
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
        "Cookie":"buvid3=2F38CD55-CCD9-0D05-EFAC-D78F4FCEE3A133631infoc; b_nut=1691060433; i-wanna-go-back=-1; _uuid=E37F628D-CE5A-5DD1-B23C-910B92326A76633722infoc; FEED_LIVE_VERSION=V8; header_theme_version=CLOSE; SESSDATA=d324dcc4%2C1706612493%2C8ce13%2A81zqyFrgt0rrTutbzOcf6NXii0x3EXBwvDIT9w6zs4rXoM6miWp779yNngwMbCD26szHztpgAAEgA; bili_jct=348a40f9dff0f5a035a9bec3dd91083c; DedeUserID=520029018; DedeUserID__ckMd5=179dfa6087c5f3f9; rpdid=|(mmJlY|~||0J'uYmu|Y|Rm); buvid4=0A6B4ED8-EFBE-C823-919F-2D38E9352F7055238-023020811-AYMpmfEzGjyejvuh2eCCkA%3D%3D; buvid_fp_plain=undefined; nostalgia_conf=-1; b_ut=5; is-2022-channel=1; LIVE_BUVID=AUTO1116911562759162; CURRENT_QUALITY=116; hit-new-style-dyn=1; hit-dyn-v2=1; CURRENT_BLACKGAP=0; fingerprint=d1f57f19105afe876875f4d406cae4a6; CURRENT_FNVAL=4048; home_feed_column=5; browser_resolution=1699-953; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2OTQxODU1MTIsImlhdCI6MTY5MzkyNjMxMiwicGx0IjotMX0.gFAVbUppg5H_wIZGERddzOAdrhwXERwn1ImjtxkE2AY; bili_ticket_expires=1694185512; PVID=3; buvid_fp=d1f57f19105afe876875f4d406cae4a6; b_lsid=12A610B5C_18A68640A2F; sid=6ocelinu; bp_video_offset_520029018=837948252620849161"
    }
    #遍历所有网页，生成总的html文件
    for page in range(math.ceil(video_num/30)):
        cur_url=url+"&page="+str(page)
        response = requests.get(cur_url,headers=header)
        html+=response.text
    return html

# 解析HTML，提取视频链接
def get_video_links(html):
    soup = BeautifulSoup(html, "html.parser")
    bvids = re.findall(r'bvid:"([^"]+)"', html)#获取视频bv号，与固定的字符串连接即可得到视频链接
    video_links = []
    for vid in bvids:
        video_links.append("https://www.bilibili.com/video/"+vid)
    return video_links

def tranfrom_url(url):
    #将视频链接转换为可获取弹幕地址的网页链接(www.ibilibili.com)
    url_index = url.find('bilibili')
    new_url = url[:url_index] + 'i' + url[url_index:]
    return new_url

# 解析视频页面，提取弹幕信息
def get_barrages_list(html):
    soup = BeautifulSoup(html, "html.parser")
    barrage_info = []
    barrages_url=re.findall('https://api.bilibili.com/x/v1/dm/list.so\?oid=\d+',html)#通过正则表达式从网页提出出弹幕api
    barrages_response=requests.get(barrages_url[0])
    barrages_response.encoding='utf-8'
    barrages_list = re.findall('<d p=".*?">(.*?)</d>', barrages_response.text)#从存储弹幕网页中爬取所有弹幕信息存储在列表中
    return barrages_list

# 统计弹幕数量，并按照数量进行排序
def count_and_sort_barrages(barrage_list):
    barrages_count = {}
    for barrage in barrage_list:
        if barrage in barrages_count:
            barrages_count[barrage] += 1
        else:
            barrages_count[barrage] = 1
    sorted_barrages = sorted(barrages_count.items(), key=lambda x: x[1], reverse=True)#对弹幕进行排序
    return sorted_barrages

# 输出数量排名前20的弹幕
def output_top_barrages(sorted_barrages):
    for i, (barrage, count) in enumerate(sorted_barrages[:barrages_num]):
        print(f"{i+1}. 弹幕: {barrage}，数量: {count}")

# 将结果保存到excel文件中
def save_excel(sorted_barrages):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value='排名'
    ws.cell(row=1, column=2).value = '数量'
    ws.cell(row=1, column=3).value = '弹幕'
    for i, row in enumerate(sorted_barrages[:20]):
        ws.cell(row=i + 2, column=1).value = 'No.'+str(i+1)
        ws.cell(row=i + 2, column=2).value = row[1]
        ws.cell(row=i + 2, column=3).value = row[0]
    wb.save('output.xlsx')

# 生成词云图
def create_wordcloud(sorted_barrages):
    barrages_text = [x[0] for x in sorted_barrages]
    stop = {'了', '的', '是', '吧', '啊', '你', '我', '他','日本','吗','就','都','这','说','不','支持'}
    barrages_cut = [word for x in barrages_text for word in jieba.lcut(x) if word not in stop]
    text = ' '.join(barrages_cut)
    background_img = np.array(Image.open('earth_mask.jpg'))
    # 生成对象
    wc = WordCloud(font_path='simsun.ttc',
                   width=800, height=600,
                   max_words=400,
                   mode="RGBA",
                   background_color='lightblue',
                   mask=background_img,
                   stopwords=stop).generate(text)

    # 显示词云图
    plt.imshow(wc, interpolation="bilinear")
    plt.axis("off")
    plt.show()

# 主函数
def main():
    print("------开始爬取，等待一段时间------")
    search_results_html = get_search_results_html()
    video_links = get_video_links(search_results_html)
    barrage_info = []
    index = 0
    for link in video_links[:video_num]:
        new_link=tranfrom_url(link)
        video_page_html = get_video_html(new_link)
        barrage_info.extend(get_barrages_list(video_page_html))
        index+=1
        print(f'已爬取{index}条视频')
    sorted_barrages = count_and_sort_barrages(barrage_info)
    output_top_barrages(sorted_barrages)
    save_excel(sorted_barrages)
    create_wordcloud(sorted_barrages)

# 执行主函数
if __name__ == "__main__":
    main()
